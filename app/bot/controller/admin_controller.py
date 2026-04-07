import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.future import select
from sqlalchemy import func
from sqlalchemy.orm import selectinload

from app.core.extensions.utils import WORKDIR

SAFE_ROWS_PER_SHEET = 950000  # 90% of Excel's limit for safety
BATCH_SIZE = 1000  # Process users in batches


async def export_users_to_excel(
    save_path: str = str(WORKDIR.parent / "media" / "xlsx"),
) -> str:
    from app.bot.models import User
    from app.core.databases.postgres import get_general_session

    os.makedirs(save_path, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"users_export_{timestamp}.xlsx"
    full_path = os.path.join(save_path, filename)

    async with get_general_session() as session:
        total_count_result = await session.execute(select(func.count(User.id)))
        total_users = total_count_result.scalar()

        sheets_needed = max(
            1, (total_users + SAFE_ROWS_PER_SHEET - 1) // SAFE_ROWS_PER_SHEET
        )

        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="2F5597", end_color="2F5597", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        data_font = Font(name="Calibri", size=10)
        data_alignment = Alignment(horizontal="left", vertical="center")

        border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3"),
        )

        light_fill = PatternFill(
            start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
        )
        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        premium_fill = PatternFill(
            start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
        )

        # Headers
        headers = [
            "ID",
            "Telegram ID",
            "Full Name",
            "Username",
            "Language",
            "Premium",
            "Last Active",
            "Referred By",
            "Active Status",
            "Balance",
            "Created At",
            "Updated At",
            "Text Downloads",
            "Voice Downloads",
            "YouTube Downloads",
            "TikTok Downloads",
            "Likee Downloads",
            "Snapchat Downloads",
            "Instagram Downloads",
            "Twitter Downloads",
        ]

        # Process users in batches
        current_sheet_idx = 1
        current_row = 2
        current_ws = None

        offset = 0
        while True:
            result = await session.execute(
                select(User)
                .options(selectinload(User.statistics))
                .offset(offset)
                .limit(BATCH_SIZE)
                .order_by(User.id)
            )
            users_batch = result.scalars().all()

            if not users_batch:
                break

            for user in users_batch:
                if current_ws is None or current_row > SAFE_ROWS_PER_SHEET:
                    sheet_name = (
                        f"Users_Page_{current_sheet_idx}"
                        if sheets_needed > 1
                        else "Users"
                    )
                    current_ws = wb.create_sheet(sheet_name)
                    current_row = 2
                    current_sheet_idx += 1

                    # Add headers
                    for col, header in enumerate(headers, 1):
                        cell = current_ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border

                # User data
                active_status = "✅ Active" if user.is_active() else "❌ Inactive"
                premium_status = "⭐ Premium" if user.is_tg_premium else "Standard"
                stats = user.statistics
                row_data = [
                    user.id,
                    user.tg_id,
                    user.full_name,
                    f"@{user.username}" if user.username else "N/A",
                    user.language_code or "N/A",
                    premium_status,
                    (
                        user.last_active.strftime("%Y-%m-%d %H:%M")
                        if user.last_active
                        else "N/A"
                    ),
                    user.referred_by or "N/A",
                    active_status,
                    user.balance if user.balance is not None else 0.0,
                    (
                        user.created_at.strftime("%Y-%m-%d %H:%M")
                        if hasattr(user, "created_at") and user.created_at
                        else "N/A"
                    ),
                    (
                        user.updated_at.strftime("%Y-%m-%d %H:%M")
                        if hasattr(user, "updated_at") and user.updated_at
                        else "N/A"
                    ),
                    stats.from_text if stats else 0,
                    stats.from_voice if stats else 0,
                    stats.from_youtube if stats else 0,
                    stats.from_tiktok if stats else 0,
                    stats.from_like if stats else 0,
                    stats.from_snapchat if stats else 0,
                    stats.from_instagram if stats else 0,
                    stats.from_twitter if stats else 0,
                ]

                # Write row
                for col, value in enumerate(row_data, 1):
                    cell = current_ws.cell(row=current_row, column=col, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border

                    # Colors
                    if user.is_tg_premium:
                        cell.fill = premium_fill
                    elif current_row % 2 == 0:
                        cell.fill = light_fill
                    else:
                        cell.fill = white_fill

                current_row += 1

            offset += BATCH_SIZE

        # Auto-adjust columns
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Set active sheet
        if wb.worksheets:
            wb.active = wb.worksheets[0]

        # Save
        wb.save(full_path)

        return full_path
